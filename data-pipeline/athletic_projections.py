#!/usr/bin/env python3
"""
athletic_projections.py — parse The Athletic's downloadable projections
workbook into the engine's `proj` shape
=========================================================================
The Athletic (Jake's model) publishes a full bottom-up projection
workbook — team play volume -> usage shares -> per-player stat lines —
as a downloadable, user-customizable .xlsx. Unlike FantasyPros' free
rankings, it gives full STAT-LEVEL projections (attempts, yards, TDs by
category) per player, on the `QB`/`RB`/`WR`/`TE`/`DST1` sheets, alongside
its own computed points/auction-value columns.

FLUID BY DESIGN, NOT A DATA FILE: this parser takes a path to whatever
copy of the workbook the caller has — never bundle a downloaded copy into
this repo. The sheet is refreshed by The Athletic on its own schedule;
every call here re-reads the file handed to it, the same discipline the
Yahoo-paste and FantasyPros-AAV-paste importers already follow for their
own "just pasted this today" inputs.

COLUMNS ARE LOOKED UP BY HEADER TEXT, not fixed indices — mirroring
`projections.py`'s PROJ_SYN table for the exact same reason: a future
year's copy of this workbook can reorder or add columns without silently
breaking this parser, so long as the header text stays recognizable.

  pip install openpyxl
  python athletic_projections.py --season 2024 --file 2024FFBProjections.xlsx
"""
from __future__ import annotations

import argparse
import json
import re

SUFFIX = re.compile(r"\b(jr|sr|ii|iii|iv|v)\b\.?", re.I)


def norm(n: str) -> str:
    """Same normalization projections.py's FantasyPros loader uses, kept
    independent rather than imported — this module has no other dependency
    on projections.py and matching should not silently drift if that
    module's own norm() ever changes for FantasyPros-specific reasons."""
    n = (n or "").lower()
    n = re.sub(r"[.'`’]", "", n)
    n = SUFFIX.sub("", n)
    n = re.sub(r"[^a-z ]", " ", n)
    return re.sub(r"\s+", " ", n).strip()


# Header synonyms -> engine `proj` fields, same shape PROJ_SYN in
# projections.py already established (passYd/passTD/int/rushYd/rushTD/
# rec/recYd/recTD — exactly what points() in engine-core.js /
# projection_model.py expects). Header text is matched case-insensitively
# and exactly against the workbook's own abbreviations (PAYD, PATD, ...),
# observed directly from the file rather than guessed.
COL_SYN = {
    "passYd": ["payd"],
    "passTD": ["patd"],
    "int":    ["int"],
    "rushYd": ["ruyd"],
    "rushTD": ["rutd"],
    "rec":    ["rec"],
    "recYd":  ["rcyd"],
    "recTD":  ["rctd"],
}

# Sheets that carry per-player stat-line projections. DST1 has no stat
# columns (just a Custom points total) and K is absent from the workbook
# entirely — both are already pure-model (EXPERT_BLEND_W = 1.0) upstream,
# untouched, same as the FantasyPros blend's own scope.
POS_SHEETS = ["QB", "RB", "WR", "TE"]


def _header_map(row) -> dict:
    """{lowercased header text: 0-based column index}."""
    out = {}
    for i, v in enumerate(row):
        if v is not None:
            out[str(v).strip().lower()] = i
    return out


def load_parsed_json(path: str) -> dict:
    """Load a table already written by main()'s --out — the derived-data
    form committed as a validation fixture (data-pipeline/fixtures/) so a
    backtest can run in CI without the raw .xlsx (paid-subscription
    content) ever entering the repo. Same {(norm_name, pos): proj_dict}
    shape parse_workbook() returns."""
    with open(path) as f:
        payload = json.load(f)
    out = {}
    for key, rec in payload["players"].items():
        name, pos = key.rsplit("|", 1)
        out[(name, pos)] = rec
    return out


def parse_workbook(path: str) -> dict:
    """{(norm_name, pos): proj_dict} across every position sheet.

    proj_dict has whatever of passYd/passTD/int/rushYd/rushTD/rec/recYd/
    recTD the sheet actually carries for that position (QB has no
    rec/recYd/recTD columns; WR/TE have no passing columns; RB/WR have a
    RUYD/RUTD pair too — a runningback's or a receiver's OWN rushing
    stats, which is exactly what points()'s rushYd/rushTD already score).
    Also carries "_team" for the caller's own name+team matching.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out: dict = {}
    for sheet in POS_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        header = _header_map(next(rows))
        name_c = header.get("player")
        team_c = header.get("tm")
        if name_c is None:
            continue
        col_for = {}
        for field, syns in COL_SYN.items():
            for syn in syns:
                if syn in header:
                    col_for[field] = header[syn]
                    break
        for row in rows:
            if row is None or name_c >= len(row):
                continue
            name = row[name_c]
            if not name or not isinstance(name, str):
                continue
            proj = {}
            for field, idx in col_for.items():
                v = row[idx] if idx < len(row) else None
                if v is not None:
                    proj[field] = float(v)
            team = row[team_c] if team_c is not None and team_c < len(row) else None
            key = (norm(name), sheet)
            out[key] = {**proj, "_name": name, "_team": team}
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True, help="path to the .xlsx workbook")
    ap.add_argument("--season", type=int, required=True, help="season this copy projects (metadata only)")
    ap.add_argument("--out", help="write the parsed {name|pos: proj} table as JSON")
    args = ap.parse_args()

    parsed = parse_workbook(args.file)
    print(f"season {args.season}: parsed {len(parsed)} players across {POS_SHEETS} from {args.file}")
    by_pos = {}
    for (_, pos), _ in parsed.items():
        by_pos[pos] = by_pos.get(pos, 0) + 1
    print(f"  by position: {by_pos}")

    if args.out:
        serializable = {f"{k[0]}|{k[1]}": v for k, v in parsed.items()}
        with open(args.out, "w") as f:
            json.dump({"season": args.season, "players": serializable}, f, indent=2)
        print(f"  wrote {args.out}")


if __name__ == "__main__":
    main()
