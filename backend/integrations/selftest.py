"""
integrations/selftest.py
=======================
Fixture tests for the deterministic core of the league importers: player
matching and the ESPN/Yahoo settings+roster translation. These run without any
network or database (the live provider fetches must be validated against prod),
so they are the regression guard for the parsing/mapping logic.

Run:  python -m integrations.selftest   (from the backend/ dir)
"""
from __future__ import annotations

import json

from .base import (NormLeague, NormPlayer, NormTeam, opponent_team_ids, resolve_my_team,
                   resolve_my_team_index, resolve_opponent_index)
from .matching import build_index, match_player, keeper_candidates
from . import espn, espn_draft_ws, live, yahoo, yahoo_paste
from . import fantasypros_aav_paste as aav_paste
from . import athletic_upload
from . import scoring_paste


def test_matching():
    rows = [
        {"id": 1, "name": "Patrick Mahomes", "pos": "QB", "team": "KC"},
        {"id": 2, "name": "A.J. Brown", "pos": "WR", "team": "PHI"},
        {"id": 3, "name": "Michael Pittman Jr.", "pos": "WR", "team": "IND"},
        {"id": 4, "name": "Kenneth Walker III", "pos": "RB", "team": "SEA"},
        {"id": 5, "name": "Marvin Harrison", "pos": "WR", "team": "ARI"},
        {"id": 6, "name": "Marvin Harrison", "pos": "WR", "team": "IND"},
        {"id": 7, "name": "49ers D/ST", "pos": "DST", "team": "SF"},
        {"id": 8, "name": "Justin Jefferson", "pos": "WR", "team": "MIN"},
    ]
    idx = build_index(rows)
    cases = [
        (NormPlayer("Patrick Mahomes", "QB", "KC"), 1),
        (NormPlayer("AJ Brown", "WR", "PHI"), 2),
        (NormPlayer("Michael Pittman", "WR", "IND"), 3),
        (NormPlayer("Kenneth Walker", "RB", "SEA"), 4),
        (NormPlayer("Marvin Harrison", "WR", "ARI"), 5),
        (NormPlayer("Marvin Harrison", "WR", "IND"), 6),
        (NormPlayer("Marvin Harrison", "WR", "ZZZ"), None),   # ambiguous, no team
        (NormPlayer("San Francisco 49ers", "DST", "SF"), 7),   # DST by team abbrev
        (NormPlayer("Niners D/ST", "DST", "SF"), 7),           # different name, same team -> ok
        (NormPlayer("San Francisco 49ers", "DST", "ZZZ"), None),  # wrong team -> no match
        (NormPlayer("Justin Jefferson", "RB", "MIN"), 8),      # wrong pos, unique name
        (NormPlayer("Nobody Here", "WR", "KC"), None),
    ]
    for np, exp in cases:
        got = match_player(idx, np)
        assert got == exp, f"match {np.name}/{np.pos}/{np.team}: got {got}, expected {exp}"


def test_matching_nicknames():
    """The platform prints a nickname where the pool holds the legal name.

    Previously imported as an unmatched player, which is how a roster comes
    back with holes in it. The alias tier is deliberately the weakest one: it
    resolves a unique candidate or an outright team agreement, and gives up
    rather than guessing when two real players collide.
    """
    rows = [
        {"id": 1, "name": "Joshua Palmer", "pos": "WR", "team": "LAC"},
        {"id": 2, "name": "Chigoziem Okonkwo", "pos": "TE", "team": "TEN"},
        {"id": 3, "name": "Cameron Ward", "pos": "QB", "team": "TEN"},
        # Two real players a nickname fold alone cannot separate.
        {"id": 4, "name": "Michael Thomas", "pos": "WR", "team": "NO"},
        {"id": 5, "name": "Mike Thomas", "pos": "WR", "team": "LAR"},
        # An exact name must still win over any alias reasoning.
        {"id": 6, "name": "Mike Williams", "pos": "WR", "team": "NYJ"},
    ]
    idx = build_index(rows)
    cases = [
        (NormPlayer("Josh Palmer", "WR", "LAC"), 1),
        (NormPlayer("Josh Palmer", "WR", ""), 1),        # unique alias, no team needed
        (NormPlayer("Chig Okonkwo", "TE", "TEN"), 2),
        (NormPlayer("Cam Ward", "QB", "TEN"), 3),
        (NormPlayer("Michael Thomas", "WR", "NO"), 4),   # exact match, unaffected
        (NormPlayer("Mike Thomas", "WR", "LAR"), 5),     # exact match, unaffected
        (NormPlayer("Mike Thomas", "WR", ""), 5),        # exact beats the alias tier
        (NormPlayer("Mike Williams", "WR", "NYJ"), 6),
        (NormPlayer("Michael Williams", "WR", ""), 6),   # alias resolves uniquely
        # A nickname AND a wrong position is two inferences stacked, and the
        # alias tier is deliberately position-scoped so it does not resolve.
        # Conservative on purpose: an unmatched player is reported, a wrongly
        # matched one is not.
        (NormPlayer("Josh Palmer", "RB", "LAC"), None),
    ]
    for np, exp in cases:
        got = match_player(idx, np)
        assert got == exp, f"match {np.name}/{np.pos}/{np.team}: got {got}, expected {exp}"

    # Ambiguity must NOT resolve: two aliasable rows, and the caller names a
    # team belonging to neither. A wrong id here silently drafts the wrong man.
    amb = build_index([
        {"id": 10, "name": "Michael Thomas", "pos": "WR", "team": "NO"},
        {"id": 11, "name": "Mike Thomas", "pos": "WR", "team": "LAR"},
    ])
    got = match_player(amb, NormPlayer("Mikey Thomas", "WR", "ZZZ"))
    assert got is None, f"ambiguous nickname should not match, got {got}"



def test_espn_draft_picks():
    """The draft, not just the survivors.

    parse_teams() walks END-OF-SEASON ROSTERS and hangs each player's draft
    price off them, so a player who was drafted and later cut disappears. That
    is correct for keeper eligibility and wrong for learning what the room
    PAYS: the survivors skew toward the picks that worked, because expensive
    players get held and cheap busts get dropped. This asserts the draft list
    keeps the dropped pick, and reports honestly when it cannot name him.
    """
    data = {
        "id": 7,
        "settings": {"name": "L", "size": 2,
                     "draftSettings": {"type": "AUCTION", "auctionBudget": 200},
                     "scoringSettings": {"scoringItems": []},
                     "rosterSettings": {"lineupSlotCounts": {}}},
        "teams": [
            {"id": 1, "name": "Alpha", "roster": {"entries": [
                {"playerPoolEntry": {"player": {"id": 11, "fullName": "Kept Back",
                                                "defaultPositionId": 2, "proTeamId": 12}}},
            ]}},
            {"id": 2, "name": "Beta", "roster": {"entries": [
                # Acquired on waivers: on the roster, never drafted.
                {"playerPoolEntry": {"player": {"id": 99, "fullName": "Waiver Add",
                                                "defaultPositionId": 3, "proTeamId": 25}}},
            ]}},
        ],
        "draftDetail": {"picks": [
            {"teamId": 1, "playerId": 11, "bidAmount": 55, "roundId": 1},
            # Drafted for $3 by Beta and cut mid-season: on NO roster.
            {"teamId": 2, "playerId": 42, "bidAmount": 3, "roundId": 9},
        ]},
    }

    picks = espn.parse_draft_picks(data)
    assert len(picks) == 2, f"draft list must keep every pick, got {len(picks)}"
    by_id = {p.ext_id: p for p in picks}

    kept = by_id["11"]
    assert kept.pos == "RB" and kept.bid == 55 and kept.resolved
    assert kept.owner == "Alpha"

    dropped = by_id["42"]
    assert dropped.bid == 3, "the dropped player's PRICE survives in the draft data"
    assert dropped.pos == "" and not dropped.resolved, "…but ESPN's draft gives no position"
    assert dropped.owner == "Beta", "owner comes from who drafted him"

    # The waiver add is on a roster but was never drafted -> not a draft pick.
    assert "99" not in by_id

    # Unresolved ids are reported so the caller can go and look them up.
    assert espn.unresolved_pick_ids(picks) == [42]

    # With a player-info lookup supplied, the dropped pick is fully named.
    lookup = espn.parse_player_info({"players": [
        {"player": {"id": 42, "fullName": "Cut Wideout", "defaultPositionId": 3, "proTeamId": 25}},
    ]})
    assert lookup[42]["pos"] == "WR" and lookup[42]["team"] == "SF"
    resolved = {p.ext_id: p for p in espn.parse_draft_picks(data, lookup)}["42"]
    assert resolved.pos == "WR" and resolved.name == "Cut Wideout" and resolved.resolved

    # parse_player_info accepts the bare-list shape ESPN also returns.
    assert espn.parse_player_info([
        {"player": {"id": 5, "fullName": "X", "defaultPositionId": 1, "proTeamId": 12}}
    ])[5]["pos"] == "QB"

    # And the whole league carries the draft alongside the rosters.
    lg = espn.parse_league(data, season=2025, my_team="Alpha", pos_by_id=lookup)
    assert len(lg.draft_picks) == 2
    assert len(lg.teams[0].players) == 1, "rosters are unchanged by any of this"
    priced = [p for p in lg.draft_picks if p.bid]
    assert sum(p.bid for p in priced) == 58, "full draft spend, including the dropped $3"


def test_espn():
    data = {
        "id": 123456,
        "settings": {
            "name": "Dynasty Warriors", "size": 12,
            "scoringSettings": {"scoringItems": [{"statId": 53, "points": 0.5}]},
            "draftSettings": {"type": "AUCTION", "auctionBudget": 300},
            "rosterSettings": {"lineupSlotCounts": {
                "0": 1, "2": 2, "4": 2, "6": 1, "23": 1, "7": 1, "17": 1, "16": 1, "20": 6, "21": 1}},
        },
        "teams": [
            {"id": 1, "name": "Team Ari", "roster": {"entries": [
                {"playerPoolEntry": {"player": {"id": 11, "fullName": "Patrick Mahomes", "defaultPositionId": 1, "proTeamId": 12}}},
            ]}},
            {"id": 2, "location": "Gridiron", "nickname": "Gurus", "roster": {"entries": [
                {"playerPoolEntry": {"player": {"id": 1, "fullName": "San Francisco 49ers", "defaultPositionId": 16, "proTeamId": 25}}},
            ]}},
        ],
        "draftDetail": {"picks": [{"teamId": 1, "playerId": 11, "bidAmount": 55, "roundId": 3}]},
    }
    lg = espn.parse_league(data, season=2026, my_team="Team Ari")
    assert lg.fmt == "auction" and lg.settings["budget"] == 300
    assert lg.settings["ppr"] == 0.5 and lg.settings["teams"] == 12
    assert lg.settings["superflex"] is True and lg.settings["roster"]["SF"] == 1
    assert lg.settings["roster"]["FLEX"] == 1 and lg.settings["roster"]["BENCH"] == 6
    assert lg.teams[0].is_mine and lg.teams[0].players[0].bid == 55
    assert lg.teams[0].players[0].round == 3
    assert lg.teams[1].players[0].pos == "DST" and lg.teams[1].players[0].team == "SF"
    # DST was not in the draft -> no bid/round (a free-agent keeper).
    assert lg.teams[1].players[0].bid is None and lg.teams[1].players[0].round is None


def test_opponent_team_ids():
    """Real team names -> settings.opponents + a name->team_id lookup, so an
    imported opponent's picks attribute to the right label instead of an
    'Unassigned' bucket or generic 'Team N' placeholder. Also returns each
    opponent's platform team id (ext_id), index-aligned with names — the
    STABLE key `resolve_opponent_index` prefers over matching by name during
    live-draft sync (see `test_sync_draft_opponent_rename`)."""
    teams = [
        NormTeam(name="Me", is_mine=True, ext_id="1"),
        NormTeam(name="The Gridiron Gurus", is_mine=False, ext_id="2"),
        NormTeam(name="Dynasty Warriors", is_mine=False, ext_id="3"),
    ]
    names, by_name, ext_ids = opponent_team_ids(teams)
    assert names == ["The Gridiron Gurus", "Dynasty Warriors"], names
    assert by_name == {"The Gridiron Gurus": 0, "Dynasty Warriors": 1}, by_name
    assert ext_ids == ["2", "3"], ext_ids
    # "my" team never appears in opponents / gets no team_id.
    assert "Me" not in by_name

    # A name clash keeps the first team's index (stable, doesn't crash).
    clash = [NormTeam(name="Team A", is_mine=False), NormTeam(name="Team A", is_mine=False)]
    names2, by_name2, ext_ids2 = opponent_team_ids(clash)
    assert names2 == ["Team A", "Team A"]
    assert by_name2 == {"Team A": 0}
    assert ext_ids2 == [None, None]

    # No opponents (single-team fixture / everyone unnamed) -> empty, not an error.
    assert opponent_team_ids([NormTeam(name="Me", is_mine=True)]) == ([], {}, [])


def test_resolve_my_team():
    """The tiered "which of these is MY team" match.

    Regression: this used to be exact-only (`key in (id, name.lower())`), so a
    name typed with different punctuation/spacing — or not typed at all —
    marked NO team as mine. `opponent_team_ids` then excluded nobody, and a
    14-team league imported 14 opponents instead of 13; both draft rooms
    render "You" PLUS all 14, with the user's own team sitting there as a
    rival that never drafts anyone. Hit in practice.
    """
    teams = [
        NormTeam(name="Team 1", ext_id="1"),
        NormTeam(name="Ari's Astounding Team", ext_id="8"),
        NormTeam(name="andrew's Angry Team", ext_id="9"),
    ]

    # Tier 1 — exact display name, and exact platform id.
    assert resolve_my_team(teams, "Ari's Astounding Team") == 1
    assert resolve_my_team(teams, "8") == 1
    # Case is not significant at any tier.
    assert resolve_my_team(teams, "ARI'S ASTOUNDING TEAM") == 1

    # Tier 2 — punctuation and spacing folded away. This is the shape that
    # actually broke: an apostrophe retyped as a different character, or lost.
    assert resolve_my_team(teams, "Aris Astounding Team") == 1
    assert resolve_my_team(teams, "  ari's   astounding team  ") == 1

    # Tier 3 — unique substring, either direction.
    assert resolve_my_team(teams, "Ari's Astounding") == 1
    assert resolve_my_team(teams, "Ari's Astounding Team 2026") == 1

    # Refusals. Each of these must return None rather than guess — attributing
    # MY picks to a rival is worse than a visibly wrong team count, which both
    # rooms now warn about.
    assert resolve_my_team(teams, "") is None
    assert resolve_my_team(teams, None) is None
    assert resolve_my_team(teams, "Nobody's Team") is None
    # Ambiguous substring: "team" is inside all three names -> give up.
    assert resolve_my_team(teams, "Team") is None

    # And the whole point: when a team IS resolved, it drops out of opponents.
    for i, t in enumerate(teams):
        t.is_mine = (i == 1)
    names, _, _ = opponent_team_ids(teams)
    assert names == ["Team 1", "andrew's Angry Team"], names


def test_sync_draft_opponent_rename():
    """`main.py sync_draft`'s own opponent-attribution lookup — a FOURTH
    instance of the exact-only-match bug class `test_resolve_my_team`/
    `test_resolve_team_ids` already regression-test twice over, found while
    answering "what happens if someone renames their team right before the
    draft?". `settings.opponents` is a snapshot frozen at IMPORT time; a
    live-sync pick's `owner` (here modeled as `lp.owner`) comes from a FRESH
    poll of the platform's live draft data. A team renamed on the platform
    between import and draft day made those two strings differ, and the
    original `{name: i for i, name in enumerate(opponents)}` dict lookup
    returned None silently — the pick still logged (`mine=False`, off the
    board) but its `team_id` never resolved, the identical "shows as sold but
    unassigned" bug already fixed twice over for keepers.

    First fix: route through `resolve_my_team_index`'s tiered name match.
    Real improvement, but still name-based — a rename severe enough to defeat
    every tier (folding punctuation, a unique substring) still loses. Asked
    directly afterward: "map team names to team ids throughout and use that
    as the constant? that would survive a serious name change." Correct, and
    ESPN/Yahoo already hand over a stable platform team id (teamId /
    team_key) with every live pick — it just wasn't being carried past
    `LivePick.owner` into anything that could use it. `resolve_opponent_index`
    tries that id FIRST, over `opponent_pairs = list(zip(opponentIds,
    opponents))` — the exact call shape `main.py` now uses per-pick — and
    only falls back to the name tiers when no id is available on either side.
    """
    opponents = ["Team 1", "Ari's Astounding Team", "andrew's Angry Team"]
    ext_ids = ["101", "102", "103"]
    pairs = list(zip(ext_ids, opponents))

    # Untouched name -> exact match, same as before (no ext_id offered here).
    assert resolve_opponent_index(pairs, None, "Ari's Astounding Team") == 1

    # Renamed on the platform since import (apostrophe dropped/retyped) ->
    # falls back to the tiered name match and still finds it.
    assert resolve_opponent_index(pairs, None, "Aris Astounding Team") == 1
    assert resolve_opponent_index(pairs, None, "ARI'S ASTOUNDING TEAM") == 1

    # THE serious rename: the new name shares nothing with the old one, so
    # EVERY name tier fails — this is exactly the case the first fix could
    # not recover from. The platform's own team id hasn't changed, though,
    # and resolving on it survives a rename of any size.
    assert resolve_opponent_index(pairs, "102", "The Completely Renamed Squad") == 1

    # No ext_id available anywhere (an older league imported before this was
    # captured) and the name is unrecognizable -> still refuses rather than
    # guessing, so the pick logs with team_id=None (visibly unassigned)
    # instead of landing on the wrong roster.
    assert resolve_opponent_index(pairs, None, "The Renamed Squad") is None
    no_ids_pairs = [(None, name) for name in opponents]
    assert resolve_opponent_index(no_ids_pairs, "102", "The Renamed Squad") is None


def test_resolve_team_ids():
    """`espn.resolve_team_ids` — the live-WS-watcher/live-ingest path's OWN
    team-id lookup. A THIRD, UNMIGRATED copy of the exact-only bug
    `test_resolve_my_team` above already regression-tests twice over
    (`resolve_my_team`, `parse_live_draft`) — this one was missed because it
    lives in `espn.py` itself rather than `base.py`, with its own inline
    `mine_key in (id, name.lower())` check instead of sharing
    `resolve_my_team_index`.

    Reported live, on a real ESPN MOCK DRAFT: "it identified the other teams
    but not mine (probably name vs number for the mock)." A mock draft's
    default team has no custom name at all — `_team_name()` falls back to
    `f"Team {id}"` — so a stored display name that doesn't byte-for-byte
    match that generic label (retyped, differently punctuated, or simply
    the league's REAL name from a prior season) failed the old exact-only
    check. `teams_by_id` was never the broken half — it's built from every
    team unconditionally — only `my_team_id` came back None, so every SOLD
    event's `is_mine` was False and none of the user's own picks were ever
    credited to their roster while opponents looked completely normal.
    """
    data = {"teams": [
        {"id": 1},                                  # mock draft default: no name set
        {"id": 2},
        {"id": 3, "name": "Ari's Team"},
    ]}

    # Every team is always in teams_by_id, named or not — confirms the
    # "other teams look fine" half of the report was never in question.
    teams_by_id, mine = espn.resolve_team_ids(data, "Team 2")
    assert teams_by_id == {1: "Team 1", 2: "Team 2", 3: "Ari's Team"}, teams_by_id
    assert mine == 2, mine

    # Tier 1 — exact numeric id as a string.
    _, mine = espn.resolve_team_ids(data, "3")
    assert mine == 3, mine

    # Tier 2 — the actual reported shape: punctuation/spacing folded away.
    _, mine = espn.resolve_team_ids(data, "Aris Team")
    assert mine == 3, mine

    # No match -> None, not a silent wrong guess (same refusal discipline as
    # resolve_my_team_index itself).
    _, mine = espn.resolve_team_ids(data, "Nobody's Team")
    assert mine is None, mine
    _, mine = espn.resolve_team_ids(data, None)
    assert mine is None, mine


def test_scoring_diagnostics():
    """ESPN/Yahoo: only PPR is auto-mapped into valuations (statId/stat_id
    schemes for the rest are undocumented and not guessed — see the docstrings
    on raw_scoring_items/raw_stat_modifiers) but the FULL raw scoring rules are
    still surfaced, not silently dropped, so the import report can point at
    League Settings -> Scoring instead of pretending everything was detected."""
    espn_data = {
        "settings": {"scoringSettings": {"scoringItems": [
            {"statId": 53, "points": 0.5},                          # PPR (mapped)
            {"statId": 4, "points": 6, "isReverseItem": False},     # unmapped, still surfaced
            {"statId": 20, "points": -2, "isReverseItem": True},    # unmapped, still surfaced
        ]}}
    }
    raw = espn.raw_scoring_items(espn_data)
    assert len(raw) == 3
    assert {"statId": 53, "points": 0.5, "isReverseItem": None} in raw
    assert {"statId": 4, "points": 6, "isReverseItem": False} in raw

    yahoo_node = [
        {"name": "L"},
        {"settings": [{"stat_modifiers": {"stats": [
            {"stat": {"stat_id": "11", "value": "1"}},   # PPR (mapped)
            {"stat": {"stat_id": "2", "value": "6"}},    # unmapped, still surfaced
        ]}}]},
    ]
    raw_y = yahoo.raw_stat_modifiers(yahoo_node)
    assert raw_y == [{"stat_id": "11", "value": "1"}, {"stat_id": "2", "value": "6"}], raw_y


def test_keeper_candidates():
    # Current-season pool the keepers must map onto.
    idx = build_index([
        {"id": 11, "name": "Patrick Mahomes", "pos": "QB", "team": "KC"},
        {"id": 7, "name": "49ers D/ST", "pos": "DST", "team": "SF"},
    ])
    data = {
        "id": 123456,
        "settings": {"name": "L", "size": 12,
                     "scoringSettings": {"scoringItems": []},
                     "draftSettings": {"type": "AUCTION", "auctionBudget": 200},
                     "rosterSettings": {"lineupSlotCounts": {"0": 1, "20": 6}}},
        "teams": [
            {"id": 1, "name": "Team Ari", "roster": {"entries": [
                {"playerPoolEntry": {"player": {"id": 11, "fullName": "Patrick Mahomes", "defaultPositionId": 1, "proTeamId": 12}}},
            ]}},
            {"id": 2, "name": "Rivals", "roster": {"entries": [
                {"playerPoolEntry": {"player": {"id": 1, "fullName": "San Francisco 49ers", "defaultPositionId": 16, "proTeamId": 25}}},
            ]}},
        ],
        "draftDetail": {"picks": [{"teamId": 1, "playerId": 11, "bidAmount": 55, "roundId": 3}]},
        # ESPN football serves waiver claims through the ACTIVITY feed: topics ->
        # messages, messageTypeId 180, targetId=player, from=winning FAAB bid.
        "topics": [
            {"messages": [{"messageTypeId": 180, "targetId": 11, "to": 1, "from": 20}]},
            {"messages": [{"messageTypeId": 180, "targetId": 11, "to": 1, "from": 35}]},
            {"messages": [{"messageTypeId": 180, "targetId": 1, "to": 2, "from": 12}]},
            # Ignored: a plain FA add (178, no bid) and a $0 priority claim.
            {"messages": [{"messageTypeId": 178, "targetId": 11, "to": 1, "from": 99}]},
            {"messages": [{"messageTypeId": 180, "targetId": 1, "to": 2, "from": 0}]},
        ],
    }
    norm = espn.parse_league(data, season=2026, my_team="Team Ari")
    cands = keeper_candidates(norm, idx)
    assert len(cands) == 2
    me = next(c for c in cands if c["name"] == "Patrick Mahomes")
    assert me["player_id"] == 11 and me["owner"] == "Me" and me["is_mine"]
    assert me["bid"] == 55 and me["round"] == 3 and me["matched"]
    assert me["waiver"] == 35, f"Mahomes waiver should be max($20,$35)=35, got {me['waiver']}"
    fa = next(c for c in cands if c["pos"] == "DST")
    assert fa["player_id"] == 7 and fa["owner"] == "Rivals" and not fa["is_mine"]
    assert fa["bid"] is None and fa["round"] is None  # undrafted -> FA basis
    assert fa["waiver"] == 12, f"49ers waiver should be $12 (the $0 claim ignored), got {fa['waiver']}"

    # The legacy `transactions` array shape still parses, and both sources merge
    # to the highest bid per player.
    legacy = {"transactions": [
        {"status": "EXECUTED", "bidAmount": 44, "items": [{"type": "ADD", "playerId": 11}]},
        {"status": "CANCELED", "bidAmount": 99, "items": [{"type": "ADD", "playerId": 11}]},
        {"status": "EXECUTED", "bidAmount": 60, "items": [{"type": "DROP", "playerId": 11}]},
    ]}
    assert espn.all_waivers(legacy) == {11: 44}, espn.all_waivers(legacy)
    both = {**legacy, "topics": [{"messages": [{"messageTypeId": 180, "targetId": 11, "from": 50}]}]}
    assert espn.all_waivers(both) == {11: 50}, "merge should keep the higher bid"

    # The activity view is only valid on the league's /communication/ sub-resource;
    # on the base league URL ESPN 400s regardless of filter.
    for url in (espn.activity_url("1", 2025), espn.activity_url("1", 2025, history=True)):
        assert "/communication/" in url, f"activity URL missing /communication/: {url}"
        assert espn.ACTIVITY_VIEW in url, url
    # ESPN 400s this feed above 25 per page.
    assert espn.ACTIVITY_PAGE <= 25, "activity feed limit must stay <= 25"
    # Waiver history is scoped to a scoringPeriodId (week) and filtered on
    # WAIVER + WAIVER_ERROR — omit either and ESPN returns no transactions.
    u = espn.site_transactions_url("1", 2025, scoring_period=15)
    assert "scoringPeriodId=15" in u, u
    assert "view=mTransactions2" in u, u
    assert json.loads(espn.WAIVER_TXN_FILTER)["transactions"]["filterType"]["value"] == \
        ["WAIVER", "WAIVER_ERROR"], espn.WAIVER_TXN_FILTER
    # ESPN rejects `limit` without a sort (FILTER_LIMIT_MISSING_SORT), so every
    # filter variant must carry one — both on /communication/ and the base URL.
    for name, filt in espn.activity_filters(25, 0) + espn.base_activity_filters(25, 0):
        assert "sortMessageDate" in filt, f"{name} filter is missing a sort"
    # Base-league filters nest under `communication` (CommunicationGroupFilterParams).
    for name, filt in espn.base_activity_filters(25, 0):
        assert list(json.loads(filt).keys()) == ["communication"], name
        assert "topics" in json.loads(filt)["communication"], name
    # Both response shapes yield waivers.
    assert espn.all_waivers({"topics": [{"messages": [
        {"messageTypeId": 180, "targetId": 9, "from": 12}]}]}) == {9: 12}
    assert espn.all_waivers({"communication": {"topics": [{"messages": [
        {"messageTypeId": 180, "targetId": 7, "from": 33}]}]}}) == {7: 33}

    # snake, no kicker, full PPR
    snake = {"id": 9, "settings": {"size": 10,
             "scoringSettings": {"scoringItems": [{"statId": 53, "points": 1.0}]},
             "draftSettings": {"type": "SNAKE"},
             "rosterSettings": {"lineupSlotCounts": {"0": 1, "2": 2, "4": 3, "6": 1, "23": 2, "20": 5}}},
             "teams": []}
    lg2 = espn.parse_league(snake, 2026)
    assert lg2.fmt == "snake" and lg2.settings["ppr"] == 1.0
    assert lg2.settings["roster"]["K"] == 0 and lg2.settings["roster"]["WR"] == 3 and lg2.settings["roster"]["FLEX"] == 2


def test_yahoo():
    league_node = [
        {"league_key": "nfl.l.123", "name": "Yahoo Ballers", "num_teams": 12},
        {"settings": [{
            "is_auction_draft": "1", "auction_budget": "260",
            "stat_modifiers": {"stats": [{"stat": {"stat_id": "11", "value": "0.5"}}]},
            "roster_positions": [
                {"roster_position": {"position": "QB", "count": 1}},
                {"roster_position": {"position": "RB", "count": 2}},
                {"roster_position": {"position": "WR", "count": 2}},
                {"roster_position": {"position": "TE", "count": 1}},
                {"roster_position": {"position": "W/R/T", "count": 1}},
                {"roster_position": {"position": "Q/W/R/T", "count": 1}},
                {"roster_position": {"position": "K", "count": 1}},
                {"roster_position": {"position": "DEF", "count": 1}},
                {"roster_position": {"position": "BN", "count": 5}},
                {"roster_position": {"position": "IR", "count": 2}},
            ]}]},
    ]
    settings, fmt, name = yahoo.parse_settings(league_node)
    assert fmt == "auction" and settings["budget"] == 260 and settings["ppr"] == 0.5
    assert settings["teams"] == 12 and settings["superflex"] is True
    assert settings["roster"]["FLEX"] == 1 and settings["roster"]["SF"] == 1 and settings["roster"]["BENCH"] == 5

    player_fields = [
        {"player_key": "nfl.p.5"}, {"player_id": "5"}, {"name": {"full": "A.J. Brown"}},
        {"editorial_team_abbr": "phi"}, {"display_position": "WR"}, {"primary_position": "WR"},
    ]
    players = {"count": 1, "0": {"player": [player_fields]}}
    roster_seg = {"roster": {"0": {"players": players}}}
    team_meta = [
        {"team_key": "nfl.l.123.t.1"}, {"team_id": "1"}, {"name": "Team Ari"},
        {"managers": {"0": {"manager": {"guid": "MEGUID"}}}},
    ]
    teams_node = {"count": 1, "0": {"team": [team_meta, roster_seg]}}
    teams = yahoo.parse_teams(teams_node, my_guid="MEGUID")
    assert teams[0].is_mine is True
    p = teams[0].players[0]
    assert p.name == "A.J. Brown" and p.pos == "WR" and p.team == "PHI"


def test_yahoo_keeper():
    """Yahoo OAuth keeper inputs: draft results, FAAB claims, keeper flags.

    Yahoo identifies a player by `player_key` in draft results and transactions
    but by `player_id` on rosters, so the fixture mixes both forms deliberately.
    """
    # ── draft results ────────────────────────────────────────────────
    draft_json = {"fantasy_content": {"league": [
        {"league_key": "449.l.82486"},
        {"draft_results": {"count": 3,
            "0": {"draft_result": {"pick": 1, "round": 1, "team_key": "449.l.82486.t.4",
                                   "player_key": "449.p.31883", "cost": "42"}},
            "1": {"draft_result": {"pick": 2, "round": 1, "team_key": "449.l.82486.t.1",
                                   "player_key": "449.p.30977"}},
            # a pick Yahoo returns with no player (forfeited/blank) must not crash
            "2": {"draft_result": {"pick": 3, "round": 1, "team_key": "449.l.82486.t.2"}},
        }},
    ]}}
    draft = yahoo.parse_draft_results(draft_json)
    assert set(draft) == {"31883", "30977"}, draft
    assert draft["31883"]["cost"] == 42 and draft["31883"]["round"] == 1
    assert draft["30977"]["cost"] is None, "snake pick carries a round, not a price"
    assert draft["30977"]["round"] == 1 and draft["30977"]["pick"] == 2

    # ── waiver / FAAB claims ─────────────────────────────────────────
    def add_txn(bid, pkey, status="successful", ptype="add"):
        return {"transaction": [
            {"transaction_key": f"449.l.1.tr.{bid}", "type": "add/drop",
             "status": status, "faab_bid": bid},
            {"players": {"count": 1, "0": {"player": [
                [{"player_key": pkey}, {"name": {"full": "X"}}],
                {"transaction_data": [{"type": ptype, "source_type": "waivers"}]},
            ]}}},
        ]}
    txn_json = {"fantasy_content": {"league": [
        {"league_key": "449.l.82486"},
        {"transactions": {"count": 5,
            "0": add_txn(17, "449.p.40000"),
            "1": add_txn(31, "449.p.40000"),            # same player, higher bid
            "2": add_txn(9, "449.p.41111"),
            "3": add_txn(99, "449.p.42222", status="failed"),   # lost claim
            "4": add_txn(50, "449.p.43333", ptype="drop"),      # the drop side
        }},
    ]}}
    waivers = yahoo.parse_transactions(txn_json)
    assert waivers["40000"] == 31, "only the top winning bid counts"
    assert waivers["41111"] == 9
    assert "42222" not in waivers, "a failed claim was never paid"
    assert "43333" not in waivers, "the dropped player wasn't bid on"

    # ── rosters + keeper flags, and folding it all together ──────────
    def player(pid, name, keeper=None):
        fields = [{"player_key": f"449.p.{pid}"}, {"player_id": pid},
                  {"name": {"full": name}}, {"editorial_team_abbr": "phi"},
                  {"display_position": "WR"}]
        if keeper is not None:
            fields.append({"is_keeper": keeper})
        return {"player": [fields]}

    teams_node = {"count": 2,
        "0": {"team": [
            [{"team_key": "449.l.82486.t.1"}, {"name": "Team Ari"},
             {"managers": {"0": {"manager": {"guid": "MEGUID"}}}}],
            {"roster": {"0": {"players": {"count": 2,
                "0": player("31883", "A.J. Brown", {"status": "1", "cost": "42", "kept": "1"}),
                "1": player("40000", "Waiver Guy"),
            }}}},
        ]},
        "1": {"team": [
            [{"team_key": "449.l.82486.t.4"}, {"name": "Rivals"}],
            {"roster": {"0": {"players": {"count": 1,
                "0": player("30977", "Snake Pick"),
            }}}},
        ]},
    }

    assert yahoo.team_names_by_key(teams_node) == {
        "449.l.82486.t.1": "Team Ari", "449.l.82486.t.4": "Rivals"}

    flags = yahoo.parse_keeper_flags(teams_node)
    assert flags["31883"]["kept"] is True and flags["31883"]["cost"] == 42
    assert "40000" not in flags, "no is_keeper block means no flag, not a false one"

    teams = yahoo.parse_teams(teams_node, my_guid="MEGUID")
    kept = yahoo.attach_keeper_inputs(teams, teams_node, draft, waivers, flags)
    assert kept == ["A.J. Brown"], kept
    mine = {p.name: p for p in teams[0].players}
    assert teams[0].is_mine is True and teams[1].is_mine is False
    assert mine["A.J. Brown"].bid == 42 and mine["A.J. Brown"].keeper_ineligible is True
    # Drafted AND claimed off waivers: both are carried, the keeper RULE picks.
    assert mine["Waiver Guy"].bid is None and mine["Waiver Guy"].waiver == 31
    assert mine["Waiver Guy"].round is None, "undrafted -> the undrafted-round rule applies"
    assert teams[1].players[0].round == 1 and teams[1].players[0].bid is None

    # Candidates come out in the shape the planner already consumes from ESPN.
    index = build_index([{"id": 7, "name": "A.J. Brown", "pos": "WR", "team": "PHI"}])
    lg = NormLeague(provider="yahoo", ext_id="449.l.82486", name="L", season=2025,
                    fmt="auction", settings={}, teams=teams)
    cands = {c["name"]: c for c in keeper_candidates(lg, index)}
    assert cands["A.J. Brown"]["owner"] == "Me" and cands["A.J. Brown"]["is_mine"] is True
    assert cands["A.J. Brown"]["bid"] == 42 and cands["A.J. Brown"]["matched"] is True
    assert cands["A.J. Brown"]["keeper_ineligible"] is True
    assert cands["Snake Pick"]["owner"] == "Rivals" and cands["Snake Pick"]["round"] == 1
    assert cands["Waiver Guy"]["waiver"] == 31 and cands["Waiver Guy"]["matched"] is False

    # Missing/absent transactions must degrade to draft-only, never explode.
    assert yahoo.parse_transactions({}) == {}
    assert yahoo.parse_draft_results({}) == {}
    assert yahoo.parse_keeper_flags({}) == {}


def test_live_draft():
    """Live sync joins the DRAFT endpoint (order/owner/price) to the ROSTER
    endpoint (names) on both platforms, because neither alone identifies who
    was taken. Mid-draft the two views disagree briefly — a pick whose player
    hasn't hit a roster yet must be skipped, not logged as an unknown."""

    # ── ESPN ─────────────────────────────────────────────────────────
    espn_data = {
        "id": 42,
        "settings": {"name": "L", "size": 2,
                     "scoringSettings": {"scoringItems": []},
                     "draftSettings": {"type": "SNAKE"},
                     "rosterSettings": {"lineupSlotCounts": {"0": 1, "20": 2}}},
        "teams": [
            {"id": 1, "name": "Team Ari", "roster": {"entries": [
                {"playerPoolEntry": {"player": {"id": 11, "fullName": "Patrick Mahomes",
                                                "defaultPositionId": 1, "proTeamId": 12}}},
            ]}},
            {"id": 2, "name": "Rivals", "roster": {"entries": [
                {"playerPoolEntry": {"player": {"id": 22, "fullName": "A.J. Brown",
                                                "defaultPositionId": 3, "proTeamId": 21}}},
            ]}},
        ],
        "draftDetail": {"inProgress": True, "picks": [
            {"playerId": 22, "teamId": 2, "roundId": 1, "roundPickNumber": 1, "overallPickNumber": 1},
            {"playerId": 11, "teamId": 1, "roundId": 1, "roundPickNumber": 2, "overallPickNumber": 2},
            # just taken — not on a roster in this payload yet
            {"playerId": 99, "teamId": 2, "roundId": 2, "roundPickNumber": 1, "overallPickNumber": 3},
        ]},
    }
    st = espn.parse_live_draft(espn_data, my_team="Team Ari")
    assert [p.overall for p in st.picks] == [1, 2], "ordered, and the unresolved pick is skipped"
    assert st.picks[0].name == "A.J. Brown" and st.picks[0].owner == "Rivals"
    assert st.picks[0].is_mine is False
    assert st.picks[1].name == "Patrick Mahomes" and st.picks[1].is_mine is True
    assert st.fmt == "snake" and st.meta["in_progress"] is True
    assert st.meta["drafted"] == 3 and st.meta["resolved"] == 2
    # Two contiguous picks are in, so pick 3 is on the clock.
    assert st.complete_through == 2

    # The LIVE path resolves "my team" through the same tiered matcher the
    # import path uses. It used to keep its own exact-only copy, so a name
    # typed with different punctuation or spacing matched nothing, `mine_ids`
    # stayed empty, and EVERY pick came back is_mine=False — reported from a
    # real mock draft as "trouble assigning teams".
    for typed in ("team ari", "  Team   Ari ", "Team Ari's", "Ari"):
        loose = espn.parse_live_draft(espn_data, my_team=typed)
        assert loose.picks[1].is_mine is True, f"live my-team match failed for {typed!r}"
        assert loose.picks[0].is_mine is False, f"live match over-claimed for {typed!r}"
    # Still refuses rather than guessing: no match, and an ambiguous one.
    assert espn.parse_live_draft(espn_data, my_team="Nobody").picks[1].is_mine is False
    assert espn.parse_live_draft(espn_data, my_team=None).picks[1].is_mine is False

    # A kona_player_info top-up (pos_by_id) resolves picks the roster view
    # hasn't caught up to — verified against a real in-progress draft where
    # the roster view had resolved NONE of the picks already made, not just
    # the "briefly" lagging one this fixture originally modeled.
    topped_up = espn.parse_live_draft(
        espn_data, my_team="Team Ari",
        pos_by_id={99: {"name": "Puka Nacua", "pos": "WR", "team": "LAR"}})
    assert [p.overall for p in topped_up.picks] == [1, 2, 3]
    assert topped_up.picks[2].name == "Puka Nacua" and topped_up.picks[2].owner == "Rivals"
    assert topped_up.meta["drafted"] == 3 and topped_up.meta["resolved"] == 3
    assert topped_up.complete_through == 3
    # Rosters still win when both know a player — pos_by_id only fills gaps.
    roster_wins = espn.parse_live_draft(
        espn_data, my_team="Team Ari",
        pos_by_id={22: {"name": "WRONG NAME", "pos": "WR", "team": "XX"}})
    assert roster_wins.picks[0].name == "A.J. Brown"

    # A live (not-yet-complete) draft can pre-populate the WHOLE season's pick
    # slots, with playerId <= 0 for ones nobody has made yet — verified against
    # a real draft that showed "160 drafted" (the full board) on pick 1. Those
    # placeholders must not count as "drafted" or get sent to the top-up.
    pending = espn.parse_live_draft({**espn_data, "draftDetail": {"inProgress": True, "picks": [
        {"playerId": 22, "teamId": 2, "overallPickNumber": 1},
        {"playerId": 11, "teamId": 1, "overallPickNumber": 2},
        {"playerId": -1, "teamId": 2, "overallPickNumber": 3},   # not picked yet
        {"playerId": 0, "teamId": 1, "overallPickNumber": 4},    # not picked yet
    ]}})
    assert pending.meta["drafted"] == 2, pending.meta["drafted"]
    assert pending.meta["raw_pick_slots"] == 4, pending.meta["raw_pick_slots"]
    assert pending.meta["resolved"] == 2

    # overallPickNumber missing -> derived from round + pick and league size
    derived = espn.parse_live_draft({**espn_data, "draftDetail": {"picks": [
        {"playerId": 11, "teamId": 1, "roundId": 2, "roundPickNumber": 2},
    ]}})
    assert derived.picks[0].overall == 4, derived.picks[0].overall   # (2-1)*2 + 2

    # An auction draft carries the price through. `settings.draftSettings.type`
    # is overridden here too — fmt now trusts that field first (see below),
    # so a self-consistent fixture needs both to agree.
    auction_settings = {**espn_data["settings"],
                        "draftSettings": {"type": "AUCTION"}}
    auc = espn.parse_live_draft({**espn_data, "settings": auction_settings, "draftDetail": {"picks": [
        {"playerId": 11, "teamId": 1, "overallPickNumber": 1, "bidAmount": 55},
    ]}})
    assert auc.fmt == "auction" and auc.picks[0].bid == 55

    # fmt must come from draftSettings.type, not from inferring off resolved
    # bid amounts — that inference silently says "snake" for an auction draft
    # whenever ZERO picks have resolved yet, which is exactly the state a
    # live auction draft is in before its first pick joins a roster (a real
    # bug this fixture pins: reported as "still says I'm in a snake draft").
    unresolved_auction = espn.parse_live_draft({**espn_data, "settings": auction_settings,
        "draftDetail": {"picks": [{"playerId": 999, "teamId": 2, "overallPickNumber": 1}]}})
    assert unresolved_auction.meta["resolved"] == 0
    assert unresolved_auction.fmt == "auction", unresolved_auction.fmt

    # ── Yahoo ────────────────────────────────────────────────────────
    def yplayer(pid, name, pos="WR"):
        return {"player": [[{"player_key": f"449.p.{pid}"}, {"player_id": pid},
                            {"name": {"full": name}}, {"editorial_team_abbr": "phi"},
                            {"display_position": pos}]]}
    teams_json = {"fantasy_content": {"league": [
        {"league_key": "449.l.1"},
        {"teams": {"count": 2,
            "0": {"team": [
                [{"team_key": "449.l.1.t.1"}, {"name": "Team Ari"},
                 {"managers": {"0": {"manager": {"guid": "MEGUID"}}}}],
                {"roster": {"0": {"players": {"count": 1, "0": yplayer("11", "Bijan Robinson", "RB")}}}},
            ]},
            "1": {"team": [
                [{"team_key": "449.l.1.t.2"}, {"name": "Rivals"}],
                {"roster": {"0": {"players": {"count": 1, "0": yplayer("22", "A.J. Brown")}}}},
            ]},
        }},
    ]}}
    draft_json = {"fantasy_content": {"league": [
        {"league_key": "449.l.1"},
        {"draft_results": {"count": 3,
            "0": {"draft_result": {"pick": 1, "round": 1, "team_key": "449.l.1.t.2",
                                   "player_key": "449.p.22"}},
            "1": {"draft_result": {"pick": 2, "round": 1, "team_key": "449.l.1.t.1",
                                   "player_key": "449.p.11"}},
            # drafted but not yet on a roster in this snapshot
            "2": {"draft_result": {"pick": 3, "round": 2, "team_key": "449.l.1.t.1",
                                   "player_key": "449.p.99"}},
        }},
    ]}}
    ys = yahoo.parse_live_draft(draft_json, teams_json, my_guid="MEGUID")
    assert [p.overall for p in ys.picks] == [1, 2], [p.overall for p in ys.picks]
    assert ys.picks[0].name == "A.J. Brown" and ys.picks[0].owner == "Rivals"
    assert ys.picks[1].name == "Bijan Robinson" and ys.picks[1].pos == "RB"
    assert ys.picks[1].owner == "Team Ari" and ys.picks[1].is_mine is True
    assert ys.meta == {"drafted": 3, "resolved": 2}, ys.meta
    assert ys.complete_through == 2

    # Empty board before the draft starts: valid, just nothing to log.
    empty = yahoo.parse_live_draft({}, teams_json, my_guid="MEGUID")
    assert empty.picks == [] and empty.complete_through == 0
    assert espn.parse_live_draft({}).picks == []

    # A gap (platform published 1 and 3 but not 2) must not claim 3 is done.
    gapped = live.LiveDraftState(picks=[live.LivePick(overall=1, name="A"),
                                        live.LivePick(overall=3, name="C")])
    assert gapped.complete_through == 1


def test_yahoo_scope_error():
    """Yahoo reports a missing Fantasy grant as a 401 that blames credentials,
    which sends you hunting a bad token or secret. It must be recognised and
    translated, not echoed."""
    body = ('{"error":{"lang":"en-US","description":"Please provide valid credentials. '
            'OAuth oauth_problem=\\"additional_authorization_required\\", realm=\\"yahooapis.com\\""}}')
    try:
        yahoo.check_fantasy_scope(401, body)
        raise AssertionError("a missing fantasy scope must be recognised")
    except yahoo.FantasyScopeError as e:
        msg = str(e)
    # The message has to name all three real causes, in order of likelihood.
    assert "Fantasy Sports (Read)" in msg, msg
    assert "DEPLOY" in msg, "Railway only applies env vars on deploy — must be called out"
    assert "Disconnect" in msg, "an already-minted token stays under-scoped"

    # Any other failure passes through untouched, so real errors aren't masked.
    yahoo.check_fantasy_scope(401, '{"error":"invalid_token"}')
    yahoo.check_fantasy_scope(500, "")
    yahoo.check_fantasy_scope(404, None)

    # Fantasy read is requested by default: leaving it opt-in produced exactly
    # the failure above, so an unset YAHOO_SCOPE must still send a scope.
    import os as _os
    saved = _os.environ.pop("YAHOO_SCOPE", None)
    _os.environ["YAHOO_CLIENT_ID"] = "cid"
    _os.environ["YAHOO_REDIRECT_URI"] = "https://example.com"
    try:
        assert "scope=fspt-r" in yahoo.authorize_url(), yahoo.authorize_url()
        _os.environ["YAHOO_SCOPE"] = "fspt-w"
        assert "scope=fspt-w" in yahoo.authorize_url()
        _os.environ["YAHOO_SCOPE"] = "-"          # explicit opt-out
        assert "scope=" not in yahoo.authorize_url()
    finally:
        _os.environ.pop("YAHOO_SCOPE", None)
        if saved is not None:
            _os.environ["YAHOO_SCOPE"] = saved
        _os.environ.pop("YAHOO_CLIENT_ID", None)
        _os.environ.pop("YAHOO_REDIRECT_URI", None)


def test_yahoo_leagues():
    data = {"fantasy_content": {"users": {"count": 1, "0": {"user": [
        {"guid": "MEGUID"},
        {"games": {"count": 2,
            "0": {"game": [{"game_key": "449", "code": "nfl", "season": "2025"},
                           {"leagues": {"count": 1, "0": {"league": [
                               {"league_key": "449.l.82486", "name": "Friends", "season": "2025", "num_teams": "12"}]}}}]},
            "1": {"game": [{"game_key": "423", "code": "nfl", "season": "2024"},
                           {"leagues": {"count": 1, "0": {"league": [
                               {"league_key": "423.l.82486", "name": "Friends", "season": "2024", "num_teams": "12"}]}}}]},
        }},
    ]}}}}
    leagues = yahoo.parse_my_leagues(data)
    assert len(leagues) == 2 and leagues[0]["season"] == 2025  # newest first
    assert leagues[0]["key"] == "449.l.82486" and leagues[1]["key"] == "423.l.82486"
    assert leagues[1]["num_teams"] == 12


def main():
    test_matching(); print("✓ matching")
    test_matching_nicknames(); print("✓ matching nicknames")
    test_espn(); print("✓ espn parse")
    test_espn_draft_picks(); print("✓ espn draft picks (incl. dropped)")
    test_opponent_team_ids(); print("✓ opponent team ids")
    test_resolve_my_team(); print("✓ resolve my team")
    test_sync_draft_opponent_rename(); print("✓ sync_draft opponent-rename attribution")
    test_resolve_team_ids(); print("✓ resolve team ids (live-WS path)")
    test_scoring_diagnostics(); print("✓ scoring diagnostics")
    test_keeper_candidates(); print("✓ keeper candidates")
    test_yahoo(); print("✓ yahoo parse")
    test_yahoo_keeper(); print("✓ yahoo keeper inputs")
    test_live_draft(); print("✓ live draft sync")
    test_yahoo_scope_error(); print("✓ yahoo scope diagnosis")
    test_yahoo_leagues(); print("✓ yahoo leagues list")
    test_waiver_weekly_fetch(); print("✓ waiver weekly fetch")
    test_yahoo_paste(); print("✓ yahoo paste import")
    test_fantasypros_aav_paste(); print("✓ fantasypros aav paste import")
    test_athletic_upload(); print("✓ athletic projections upload (second-opinion display)")
    test_espn_draft_ws(); print("✓ espn draft websocket (protocol + accumulator)")
    test_yahoo_scoring_paste(); print("✓ yahoo scoring-page paste import")
    test_espn_scoring_paste(); print("✓ espn scoring-page paste import")
    print("\nALL INTEGRATION SELFTESTS PASS")




def test_waiver_weekly_fetch():
    """fetch_league must sweep scoring periods and take the highest EXECUTED
    FAAB bid per player (ESPN returns transactions only per-week)."""
    import asyncio
    import types
    import urllib.parse as up

    league = {"id": 1, "settings": {"name": "L", "size": 12,
              "scoringSettings": {"scoringItems": []},
              "draftSettings": {"type": "AUCTION", "auctionBudget": 200},
              "rosterSettings": {"lineupSlotCounts": {"0": 1, "20": 6}}},
              "teams": [{"id": 1, "name": "T", "roster": {"entries": [
                  {"playerPoolEntry": {"player": {"id": 11, "fullName": "P M",
                                                  "defaultPositionId": 1, "proTeamId": 12}}}]}}],
              "draftDetail": {"picks": [{"teamId": 1, "playerId": 11, "bidAmount": 15, "roundId": 3}]}}

    class Resp:
        def __init__(self, code, payload): self.status_code, self._p = code, payload
        def json(self): return self._p
        def raise_for_status(self): pass

    class Fake:
        async def __aenter__(self): return self
        async def __aexit__(self, *a): return False
        async def get(self, url, headers=None):
            q = up.parse_qs(up.urlparse(url).query)
            if "mTransactions2" not in q.get("view", []):
                return Resp(200, dict(league))
            sp = q.get("scoringPeriodId", [None])[0]
            if sp is None:
                return Resp(200, dict(league))
            if int(sp) == 5:
                return Resp(200, {**league, "transactions": [
                    {"id": "a", "status": "EXECUTED", "bidAmount": 30,
                     "items": [{"type": "ADD", "playerId": 11}]}]})
            if int(sp) == 9:
                return Resp(200, {**league, "transactions": [
                    {"id": "b", "status": "EXECUTED", "bidAmount": 44,
                     "items": [{"type": "ADD", "playerId": 11}]},
                    {"id": "c", "status": "CANCELED", "bidAmount": 99,
                     "items": [{"type": "ADD", "playerId": 11}]}]})
            return Resp(200, {**league, "transactions": []})

    real = espn.httpx
    try:
        espn.httpx = types.SimpleNamespace(AsyncClient=lambda **kw: Fake())
        lg = asyncio.run(espn.fetch_league("1", 2025, my_team="T"))
    finally:
        espn.httpx = real
    p = lg.teams[0].players[0]
    assert p.bid == 15 and p.waiver == 44, (p.bid, p.waiver)   # highest executed bid
    assert lg.meta["transactions"]["waiver_players"] == 1



def test_yahoo_paste():
    """Yahoo import with NO API access: parse the Draft Results + Starting
    Rosters pages a league member can copy out of the web UI.

    Fixture mirrors the structure of a real 10-team snake league export,
    including the details that actually broke first drafts of the parser.
    """
    # Keeper badges survive copy-paste as a trailing space (draft) / blank line
    # (rosters). Built with explicit \t and trailing spaces so the fixture can't
    # be silently "cleaned" by an editor.
    draft = "\n".join([
        "Round 1",
        "1.\tCeeDee Lamb\tMcLaurin Order",
        "2.\tJahmyr Gibbs\tBecoming BEA...",          # truncated team name
        "3.\tJustin Jefferson\tLet’s Play...",   # truncated + curly apostrophe
        "Round 2",
        "1.\tJustin Fields\tLet’s Play...",
        "2.\tTrey McBride \tBecoming BEA...",         # trailing space = keeper badge
        "3.\tJosh Jacobs\tMcLaurin Order",
        "Round 3",
        "1.\tSaquon Barkley \tMcLaurin Order",        # keeper badge
        "2.\tBo Nix\tBecoming BEA...",
        "3.\tBo Nix Jr.\tBecoming BEA...",            # same round twice => traded pick
    ])
    rosters = "\n".join([
        "Starting Rosters", "Team", " ", "Week 17", " ",
        "Becoming BEARable ", "", "Pos\tPlayer",
        "QB\t", "Jahmyr Gibbs", "Jahmyr GibbsVideo Forecast", "Final W 20-17 vs TB",
        "TE\t", "Trey McBride", "Trey McBrideVideo Forecast", "", "Final L 14-37 @ Cin",
        "DEF\t", "Saints", "Saints", "- DEF", "Final W 34-26 @ Ten",
        # last player on this team: must NOT inherit the blank line that
        # precedes the next team's header (regression guard)
        "IR\t", "Bo Nix", "Bo NixVideo Forecast", "Final W 20-13 @ KC",
        "McLaurin Order ", "", "Pos\tPlayer",
        "RB\t", "Saquon Barkley", "Saquon BarkleyVideo Forecast", "", "Final W 13-12 @ Buf",
        "WR\t", "CeeDee Lamb", "CeeDee LambVideo Forecast", "Final W 30-23 @ Was",
        "Let’s Play Golf! ", "", "Pos\tPlayer",
        "QB\t", "Justin Jefferson", "Justin JeffersonVideo Forecast", "Final W 23-10 vs Det",
        "BN\t", "Rhamondre Stevenson", "Rhamondre StevensonVideo Forecast", "Final W 42-10 @ NYJ",
    ])

    # ── draft results ────────────────────────────────────────────────
    picks = yahoo_paste.parse_draft_results(draft)
    assert len(picks) == 9, len(picks)
    kept = {p.name for p in picks if p.kept}
    assert kept == {"Trey McBride", "Saquon Barkley"}, kept
    # the badge must not leak into the stored name
    assert all(p.name == p.name.strip() for p in picks)

    # ── rosters ──────────────────────────────────────────────────────
    teams = yahoo_paste.parse_rosters(rosters)
    assert [t.name for t in teams] == ["Becoming BEARable", "McLaurin Order", "Let's Play Golf!"], \
        [t.name for t in teams]
    bear = teams[0]
    assert [s.name for s in bear.players] == ["Jahmyr Gibbs", "Trey McBride", "Saints", "Bo Nix"]
    assert bear.players[1].kept is True                    # blank-line badge
    # Regression: the LAST player of a team sits right before the next team's
    # header; an earlier version flagged every one of them as kept.
    assert bear.players[3].kept is False, "last player of a team must not be flagged kept"
    assert teams[1].players[0].kept is True                # Saquon, badge
    assert teams[1].players[1].kept is False
    # slot -> position mapping (DEF -> DST; BN/IR carry no position)
    assert bear.players[2].pos == "DST"
    assert bear.players[3].pos == ""

    # ── truncated team-name resolution ───────────────────────────────
    full = [t.name for t in teams]
    assert yahoo_paste.resolve_team_name("Becoming BEA...", full) == "Becoming BEARable"
    assert yahoo_paste.resolve_team_name("Let’s Play...", full) == "Let's Play Golf!"
    assert yahoo_paste.resolve_team_name("McLaurin Order", full) == "McLaurin Order"
    # ambiguous stem returns the input unchanged so the caller can report it
    assert yahoo_paste.resolve_team_name("X...", full) == "X..."

    # ── draft slots come from ROUND 1 ONLY (trades break serpentine) ──
    slots = yahoo_paste.draft_slots(picks, full)
    assert slots == {"McLaurin Order": 1, "Becoming BEARable": 2, "Let's Play Golf!": 3}, slots

    # ── combined league ──────────────────────────────────────────────
    lg, rep = yahoo_paste.build_league(draft, rosters, my_team="Becoming BEARable")
    assert lg.provider == "yahoo-paste" and lg.fmt == "snake"
    assert [t.is_mine for t in lg.teams] == [True, False, False]
    assert lg.settings["draftSlot"] == 2
    by_name = {p.name: p for t in lg.teams for p in t.players}
    # draft round carries the keeper cost basis; badge marks ineligibility
    assert by_name["Trey McBride"].round == 2 and by_name["Trey McBride"].keeper_ineligible
    assert by_name["Saquon Barkley"].round == 3 and by_name["Saquon Barkley"].keeper_ineligible
    assert by_name["Jahmyr Gibbs"].round == 1 and not by_name["Jahmyr Gibbs"].keeper_ineligible
    # rostered but never drafted => waiver/FA pickup, no round
    assert by_name["Saints"].round is None
    assert "Rhamondre Stevenson" in rep["undrafted_on_roster"]
    # traded picks are reported, not silently trusted for ordering
    assert any("traded picks" in w for w in rep["warnings"]), rep["warnings"]
    assert any("confirm this list" in w for w in rep["warnings"]), rep["warnings"]

    # ── what /api/leagues/import-yahoo-paste persists ────────────────
    # The paste knows every team's identity AND slot; the route writes both
    # (settings.opponents + settings.teamSlots), so opponent keeper predictions
    # start from the real draft order instead of a mid-round guess.
    names, by_name, _ = opponent_team_ids(lg.teams)
    assert names == ["McLaurin Order", "Let's Play Golf!"], names   # mine excluded
    assert by_name["McLaurin Order"] == 0
    assert set(rep["draft_slots"]) == set(rep["team_names"]), "every team needs a slot"
    # Without my_team every team is an opponent — the paste can't tell which is
    # yours, so nothing is silently guessed.
    lg2, _ = yahoo_paste.build_league(draft, rosters)
    assert opponent_team_ids(lg2.teams)[0] == rep["team_names"]
    assert lg2.settings["draftSlot"] == 1

def test_fantasypros_aav_paste():
    """Real FantasyPros auction-values cheat-sheet rows (copied straight off
    the page, not synthesized) — fixture-tested the same way yahoo_paste's
    real 10-team export is, because a format assumption that only holds on a
    hand-built example is the failure mode this whole file exists to catch."""
    text = "\n".join([
        "1.\tJahmyr Gibbs (DET - RB)\t302\t$63",
        "2.\tPuka Nacua (LAR - WR)DTD\t223\t$61",              # injury tag glued on
        "62.\tAlec Pierce (IND - WR)PUP\t139\t$16",            # different tag
        "156.\tHouston Texans (HOU - DST)\t120\t$2",           # DST: franchise name
        "182.\tBrandon Aubrey (DAL - K)\t153\t$1",
        "196.\tJordan Love (GB - QB)\t257\t$0",                # $0 is a real value, not missing
        "270.\tAustin Ekeler ( - RB)\t27\t$0",                 # free agent: blank team
        "288.\tTravis Hunter (JAC - WR,CB)\t70\t$0",           # dual-listed: first pos wins
        "",                                                    # blank lines are skipped
        "not a player row at all",                             # malformed -> skipped, not fatal
    ])
    report = aav_paste.parse_aav_sheet(text)
    assert len(report.rows) == 8, len(report.rows)
    assert report.skipped == ["not a player row at all"], report.skipped

    by_name = {r.name: r for r in report.rows}
    assert by_name["Jahmyr Gibbs"].pos == "RB" and by_name["Jahmyr Gibbs"].team == "DET"
    assert by_name["Jahmyr Gibbs"].aav == 63.0
    # the injury/practice tag must not leak into the parsed name or price
    assert by_name["Puka Nacua"].aav == 61.0
    assert by_name["Alec Pierce"].aav == 16.0
    # DST: FantasyPros gives the team directly, no name-based lookup needed
    assert by_name["Houston Texans"].pos == "DST" and by_name["Houston Texans"].team == "HOU"
    # $0 is a real, meaningful floor value, not a parse failure
    assert by_name["Jordan Love"].aav == 0.0
    # a free agent has an explicit-but-empty team, not a missing field
    assert by_name["Austin Ekeler"].team == ""
    # dual-position players are filed under the FIRST position listed
    assert by_name["Travis Hunter"].pos == "WR"

    # ── feeds the SAME matcher ESPN/Yahoo import already uses ──────────
    pool = [
        {"id": 1, "name": "Jahmyr Gibbs", "pos": "RB", "team": "DET"},
        {"id": 2, "name": "Puka Nacua", "pos": "WR", "team": "LAR"},
        {"id": 3, "name": "Houston Texans", "pos": "DST", "team": "HOU"},
        {"id": 4, "name": "Austin Ekeler", "pos": "RB", "team": "LAC"},  # team drifted
        # Alec Pierce, Brandon Aubrey, Jordan Love, Travis Hunter absent —
        # exercises the unmatched path.
    ]
    index = build_index(pool)
    norm_players = aav_paste.to_norm_players(report)
    assert len(norm_players) == 8
    matched = [(np, match_player(index, np)) for np in norm_players]
    ids = {np.name: pid for np, pid in matched}
    assert ids["Jahmyr Gibbs"] == 1
    assert ids["Puka Nacua"] == 2
    assert ids["Houston Texans"] == 3          # DST matches on team, not name
    assert ids["Austin Ekeler"] == 4           # name+pos unique even with a blank/stale team
    assert ids["Alec Pierce"] is None          # not in the pool -> unmatched, not guessed
    assert ids["Jordan Love"] is None


def test_athletic_upload():
    """The Athletic's projections workbook, parsed for a SECOND-OPINION
    display only — never a valuation input (roadmap 0.1b gated blending this
    into valuePoints the way FantasyPros is blended and it failed the
    decisive check; see CLAUDE.md). Built in-memory rather than committing
    the real .xlsx (paid-subscription content) — same reasoning as the
    fixtures/ JSON extracts used for the 0.1b backtest itself — but the
    header text and stat values are drawn from a real uploaded copy, not
    invented, the same discipline `test_fantasypros_aav_paste` uses for its
    hand-copied real rows."""
    import io
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    qb = wb.create_sheet("QB")
    qb.append(["Player", "Tm", "PAYD", "PATD", "INT", "RUYD", "RUTD"])
    qb.append(["Jacoby Brissett", "ARI", 3231.86, 21.36, 7.95, 208.65, 1.11])

    rb = wb.create_sheet("RB")
    rb.append(["Player", "Tm", "RUYD", "RUTD", "REC", "RCYD", "RCTD"])
    rb.append(["J.K. Dobbins", "DEN", 735.81, 4.78, 15.26, 74.47, 0.50])

    # A sheet the app doesn't score at all (out of scope, same as K/DST in
    # EXPERT_BLEND_W) — must be silently ignored, not error.
    dst = wb.create_sheet("DST1")
    dst.append(["Player", "Tm", "Custom"])
    dst.append(["Broncos", "DEN", 120.0])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    report = athletic_upload.parse_workbook(buf)
    assert set(report.sheets_found) == {"QB", "RB"}, report.sheets_found
    assert len(report.rows) == 2, len(report.rows)

    by_name = {r.name: r for r in report.rows}
    brissett = by_name["Jacoby Brissett"]
    assert brissett.pos == "QB" and brissett.team == "ARI"
    assert brissett.proj["passYd"] == 3231.86
    assert brissett.proj["passTD"] == 21.36
    assert brissett.proj["int"] == 7.95
    assert "rec" not in brissett.proj          # QB sheet carries no receiving columns

    dobbins = by_name["J.K. Dobbins"]
    assert dobbins.pos == "RB" and dobbins.team == "DEN"
    assert dobbins.proj["rec"] == 15.26
    assert dobbins.proj["recYd"] == 74.47

    # ── feeds the SAME matcher ESPN/Yahoo/AAV-paste import already use ──
    pool = [
        {"id": 1, "name": "Jacoby Brissett", "pos": "QB", "team": "ARI"},
        {"id": 2, "name": "J.K. Dobbins", "pos": "RB", "team": "DEN"},
    ]
    index = build_index(pool)
    norm_players = athletic_upload.to_norm_players(report)
    assert len(norm_players) == 2
    ids = {np.name: match_player(index, np) for np in norm_players}
    assert ids["Jacoby Brissett"] == 1
    assert ids["J.K. Dobbins"] == 2


def test_espn_draft_ws():
    """`espn_draft_ws.parse_ws_line` against lines copied verbatim from a real
    captured HAR of ESPN's actual draft-room WebSocket (`wss://fantasydraft.
    espn.com/.../JOIN`) — the channel ESPN's own client uses for live picks,
    which `integrations/live.py`'s REST polling can never see (see CLAUDE.md).
    Real values, not synthesized, same discipline as `test_yahoo_paste` and
    `test_fantasypros_aav_paste`."""
    p = espn_draft_ws.parse_ws_line

    # Field order corrected live, not from the original HAR guess: a real
    # user cross-checked a SOLD line against ESPN's own draft room and the
    # team at THIS position (position 1, "6" here) was the actual winning
    # bidder for three consecutive picks — the position originally labeled
    # "winning" (position 3, "13" here) held values outside the league's
    # real team-id range entirely. Player id and price positions (2 and 4)
    # were independently confirmed correct and are unchanged.
    sold = p("SOLD 6 4570037 13 1 0\n")
    assert sold == {"type": "sold", "nominating_team_id": 13, "player_id": 4570037,
                     "winning_team_id": 6, "price": 1, "flag": 0}, sold

    bid = p("BID 6 4570037 1 25000 25000\n")
    assert bid == {"type": "bid", "team_id": 6, "player_id": 4570037,
                    "amount": 1, "clock_ms": 25000, "clock_ms2": 25000}, bid

    nomination = p("NOMINATION 6 25000\n")
    assert nomination == {"type": "nomination", "team_id": 6, "clock_ms": 25000}, nomination

    passed = p("PASSED 7 3046439 false\n")
    assert passed == {"type": "passed", "team_id": 7, "player_id": 3046439, "auto": False}, passed

    # CLOCK has two shapes depending on phase — between-nomination vs. active bid.
    clock_idle = p("CLOCK 3 522\n")
    assert clock_idle == {"type": "clock", "phase": 3, "clock_ms": 522}, clock_idle
    clock_bid = p("CLOCK 2 24749 6 4570037 1\n")
    assert clock_bid == {"type": "clock", "phase": 2, "clock_ms": 24749,
                          "team_id": 6, "player_id": 4570037, "amount": 1}, clock_bid

    joined = p("JOINED 7 {B32FA1C0-4AC4-4241-9C27-345AA44C4300}\n")
    assert joined == {"type": "joined", "team_id": 7,
                       "swid": "{B32FA1C0-4AC4-4241-9C27-345AA44C4300}"}, joined

    autodraft = p("AUTODRAFT 7 false\n")
    assert autodraft == {"type": "autodraft", "team_id": 7, "enabled": False}, autodraft

    token = p("TOKEN 1:550003701:7:{B32FA1C0-4AC4-4241-9C27-345AA44C4300}:-1221336021\n")
    assert token == {"type": "token",
                      "raw": "1:550003701:7:{B32FA1C0-4AC4-4241-9C27-345AA44C4300}:-1221336021"}, token

    # PING/PONG payloads are URL-encoded ("%20" -> " ") in the wire format.
    ping = p("PING PING%201787195389520\n")
    assert ping == {"type": "ping", "payload": "PING 1787195389520"}, ping
    pong = p("PONG PING%201787195389520\n")
    assert pong == {"type": "pong", "payload": "PING 1787195389520"}, pong

    # INIT's payload is a large opaque blob (undecoded — see module docstring);
    # the type still parses so a caller can at least recognize it arrived.
    init = p("INIT AAAAAQAAAAEgyGP1AAAABw==\n")
    assert init == {"type": "init"}, init

    # Unknown/garbage lines don't raise, and don't get invented data.
    assert p("") is None
    assert p("   \n") is None
    assert p("SOMETHING_FUTURE 1 2 3\n") is None
    assert p("SOLD not-a-number 1 2 3 4\n") is None   # malformed field, not a crash

    # join_url composes the exact query-string shape from the capture, given
    # an externally-supplied hash — it does not fabricate one (see docstring).
    url = espn_draft_ws.join_url("550003701", 7, "{B32FA1C0-4AC4-4241-9C27-345AA44C4300}",
                                 -1221336021, nocache=1413)
    assert url == ("wss://fantasydraft.espn.com/game-1/league-550003701/JOIN"
                    "?1=1&2=550003701&3=7&4={B32FA1C0-4AC4-4241-9C27-345AA44C4300}"
                    "&5=1:550003701:7:{B32FA1C0-4AC4-4241-9C27-345AA44C4300}:-1221336021"
                    "&6=false&7=false&8=KONA&nocache=1413"), url

    # A SECOND, independent real capture (different league, team, and
    # resulting value) — same shape, confirming join_url generalizes rather
    # than having been fit to the one example above.
    url2 = espn_draft_ws.join_url("2053165706", 9, "{B32FA1C0-4AC4-4241-9C27-345AA44C4300}",
                                  1087841463, nocache=896252)
    assert url2 == ("wss://fantasydraft.espn.com/game-1/league-2053165706/JOIN"
                     "?1=1&2=2053165706&3=9&4={B32FA1C0-4AC4-4241-9C27-345AA44C4300}"
                     "&5=1:2053165706:9:{B32FA1C0-4AC4-4241-9C27-345AA44C4300}:1087841463"
                     "&6=false&7=false&8=KONA&nocache=896252"), url2

    # draftSecurity is the REST endpoint that hands back that exact join-hash
    # value directly (see espn.fetch_draft_security) — url shape pinned
    # against the real captured request.
    assert espn.draft_security_url("2053165706", 2026, 9) == (
        "https://lm-api-reads.fantasy.espn.com/apis/v3/games/ffl/seasons/2026"
        "/segments/0/leagues/2053165706/teams/9/draftSecurity")

    # ── accumulator: SOLD events -> LivePicks, no networking ────────────
    watcher = espn_draft_ws.LiveDraftWatcher(my_team_id=7, teams_by_id={6: "Team A", 13: "Team B"},
                                             start_overall=5)
    # First SOLD: player unresolved yet -> on_event flags it for lookup.
    needs_lookup = watcher.on_event(sold)
    assert needs_lookup == 4570037, needs_lookup
    # Re-feeding the SAME unresolved player again doesn't re-flag it.
    assert watcher.on_event(sold) is None
    # Still unresolved -> not in state() yet, but it kept its slot (position 0).
    assert watcher.state().picks == []
    watcher.add_player_info({4570037: {"name": "Puka Nacua", "pos": "WR", "team": "LAR"}})
    st = watcher.state()
    assert len(st.picks) == 2, st.picks   # both queued SOLD events now resolve
    assert st.picks[0].overall == 5 and st.picks[0].name == "Puka Nacua"
    assert st.picks[0].owner == "Team A" and st.picks[0].bid == 1   # winning_team_id=6 -> "Team A"
    assert st.picks[0].is_mine is False
    assert st.fmt == "auction"   # a real price was seen
    assert st.meta["drafted"] == 2 and st.meta["resolved"] == 2

    # A second, DIFFERENT unresolved player in a fresh watcher. winning_team_id
    # is position 1 in the wire line ("1" here, see the field-order note above).
    w2 = espn_draft_ws.LiveDraftWatcher(my_team_id=1)
    assert w2.on_event(p("SOLD 1 999999 13 5 0\n")) == 999999
    assert w2.state().picks == []   # still unresolved -> not silently guessed
    w2.add_player_info({999999: {"name": "Nobody Known", "pos": "RB", "team": "XX"}})
    assert w2.state().picks[0].is_mine is True   # winning_team_id == my_team_id
    assert w2.state().picks[0].overall == 1      # default start_overall


def test_yahoo_scoring_paste():
    """Yahoo's League Settings -> Scoring page, pasted as plain text — the
    real fix for "my Yahoo import only auto-mapped PPR, 42 other scoring
    rules not auto-mapped." `raw_stat_modifiers()` only ever sees a numeric
    `stat_id` this app has no verified mapping for; Yahoo's SCORING PAGE, by
    contrast, labels every rule in plain English, so there is nothing to
    guess. This fixture is the REAL page a user pasted, verbatim."""
    text = (
        "Offense\tLeague Value\tYahoo Default Value\n"
        "Completions\nYahoo Default\n0.5\t0\n"
        "Passing Yards\nYahoo Default\n"
        "35 yards per point; 5 points at 360 yards; 5 points at 450 yards; 5 points at 600 yards"
        "\t25 yards per point\n"
        "Passing Touchdowns\nYahoo Default\n6\t4\n"
        "Interceptions\nYahoo Default\n-2\t-1\n"
        "Rushing Yards\t10 yards per point; 5 points at 150 yards; 5 points at 250 yards\n"
        "Rushing Touchdowns\t6\n"
        "Receptions\nYahoo Default\n1\t0.5\n"
        "Receiving Yards\nYahoo Default\n"
        "15 yards per point; 5 points at 150 yards; 5 points at 250 yards\t10 yards per point\n"
        "Receiving Touchdowns\t6\n"
        "Return Yards\t35 yards per point\t0\n"
        "Return Touchdowns\t6\n"
        "2-Point Conversions\t2\n"
        "Fumbles Lost\t-2\n"
        "Offensive Fumble Return TD\t6\n"
        "40+ Yard Passing Touchdowns\nYahoo Default\n0.5\t0\n"
        "40+ Yard Rushing Touchdowns\nYahoo Default\n1\t0\n"
        "40+ Yard Receiving Touchdowns\nYahoo Default\n0.5\t0\n"
        "Kickers\tLeague Value\tYahoo Default Value\n"
        "Field Goals 0-19 Yards\t3\n"
        "Field Goals 40-49 Yards\nYahoo Default\n3\t4\n"
        "Point After Attempt Made\t1\n"
        "Point After Attempt Missed\nYahoo Default\n-1\t0\n"
        "Defense/Special Teams\tLeague Value\tYahoo Default Value\n"
        "Sack\t1\n"
        "Interception\t2\n"          # DEFENSE'S "Interception" — must NOT hit ptsPerInt
        "Points Allowed 0 points\nYahoo Default\n15\t10\n"
        "Extra Point Returned\t2\n"
    )
    r = scoring_paste.parse_yahoo_scoring_page(text)

    # The 8 ScoringRules fields this app actually models, at the LEAGUE
    # value (not Yahoo's default) — bonus brackets included in the raw
    # value are ignored (base rate only), a real value materially
    # different from this app's own DEFAULT_SCORING (0.04/4/0.1/0.1).
    assert abs(r.scoring["ptsPerPassYd"] - 1 / 35) < 1e-9
    assert r.scoring["ptsPerPassTD"] == 6
    assert r.scoring["ptsPerInt"] == -2          # OFFENSE's Interceptions, not Defense's
    assert abs(r.scoring["ptsPerRushYd"] - 0.1) < 1e-9
    assert r.scoring["ptsPerRushTD"] == 6
    assert abs(r.scoring["ptsPerRecYd"] - 1 / 15) < 1e-9
    assert r.scoring["ptsPerRecTD"] == 6
    assert r.scoring["ptsPerFumble"] == -2
    assert r.ppr == 1
    assert len(r.scoring) == 8, r.scoring

    # Bonus brackets are called out, not silently dropped.
    assert any("Passing Yards" in w for w in r.warnings)
    assert any("Rushing Yards" in w for w in r.warnings)
    assert any("Receiving Yards" in w for w in r.warnings)

    # Everything this engine has no field for (Completions, Return Yards,
    # 2-Point Conversions, every Kicker/Defense row) is surfaced, not lost.
    unmapped_labels = {u["label"] for u in r.unmapped}
    assert "Completions" in unmapped_labels
    assert "Return Yards" in unmapped_labels
    assert "2-Point Conversions" in unmapped_labels
    assert "Field Goals 0-19 Yards" in unmapped_labels
    # Defense's "Interception" (+2, a takeaway) is unmapped, NOT folded into
    # ptsPerInt (-2, a QB's turnover) just because the words are similar.
    defense_int = [u for u in r.unmapped if u["label"] == "Interception"]
    assert len(defense_int) == 1 and defense_int[0]["raw"] == "2"


def test_espn_scoring_paste():
    """ESPN's League Settings -> Scoring Settings page, pasted as plain
    text. Real captured page: each rule is one line, "<label> (<CODE>)
    <value>" with no space before the value, and per-yard categories spell
    the denominator INTO the label ("Every 25 passing yards") rather than as
    a separate phrase the way Yahoo does it."""
    text = (
        "Scoring\n"
        "Passing\n"
        "Every 25 passing yards (PY25)1\n"
        "TD Pass (PTD)4\n"
        "50+ yard TD pass bonus (PTD50)1\n"
        "Interceptions Thrown (INT)-2\n"
        "2pt Passing Conversion (2PC)2\n"
        "300-399 yard passing game (P300)2\n"
        "Rushing\n"
        "Every 10 rushing yards (RY10)1\n"
        "TD Rush (RTD)6\n"
        "40+ yard TD rush bonus (RTD40)2\n"
        "Receiving\n"
        "Every 10 receiving yards (REY10)1\n"
        "TD Reception (RETD)6\n"
        "Kicking\n"
        "Each PAT Made (PAT)1\n"
        "FG Made (0-39 yards) (FG0)3\n"
        "Team Defense / Special Teams\n"
        "Each Sack (SK)0.5\n"
        "Each Interception (INT)1\n"    # DEFENSE'S — same code "INT" as Passing's, must not collide
        "0 points allowed (PA0)10\n"
        "Miscellaneous\n"
        "Kickoff Return TD (KRTD)6\n"
        "Total Fumbles Lost (FUML)-2\n"
    )
    r = scoring_paste.parse_espn_scoring_page(text)

    assert r.scoring["ptsPerPassYd"] == 1 / 25
    assert r.scoring["ptsPerPassTD"] == 4
    assert r.scoring["ptsPerInt"] == -2       # Passing's "Interceptions Thrown"
    assert r.scoring["ptsPerRushYd"] == 0.1
    assert r.scoring["ptsPerRushTD"] == 6
    assert r.scoring["ptsPerRecYd"] == 0.1
    assert r.scoring["ptsPerRecTD"] == 6
    assert r.scoring["ptsPerFumble"] == -2     # from Miscellaneous, not Passing/Rushing/Receiving
    assert r.ppr is None                       # this page never carries PPR; ESPN's own statId-53 owns it
    assert len(r.scoring) == 8, r.scoring

    # Same-CODE collision across sections resolves by SECTION, not by the
    # "(INT)" code the two rows happen to share.
    defense_int = [u for u in r.unmapped if u["label"] == "Each Interception"]
    assert len(defense_int) == 1 and defense_int[0]["raw"] == "1"

    unmapped_labels = {u["label"] for u in r.unmapped}
    assert "50+ yard TD pass bonus" in unmapped_labels
    assert "2pt Passing Conversion" in unmapped_labels
    assert "Each PAT Made" in unmapped_labels
    assert "0 points allowed" in unmapped_labels
    assert "Kickoff Return TD" in unmapped_labels


if __name__ == "__main__":
    main()
