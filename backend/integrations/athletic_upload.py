"""
integrations/athletic_upload.py
================================
The Athletic's downloadable fantasy projections workbook, uploaded live for a
SECOND OPINION display — not a valuation input.

Background: roadmap 0.1b tried blending this same source into `valuePoints`
as a second expert alongside FantasyPros (the model roadmap 0.1 already
ships), the same way the app already treats FantasyPros. It was gated the
same two ways every signal in this codebase is required to clear — solo vs.
the pure model, then the decisive check on top of the ALREADY-SHIPPED board —
and FAILED the second one: QB/TE flip sign between the two validation
seasons, RB/WR are directionally consistent but under 0.01 Spearman, an
order of magnitude below the gains that justified shipping FantasyPros in
the first place. See CLAUDE.md "Second expert source: The Athletic" and
docs/ROADMAP.md 0.1b for the full numbers. NOTHING here feeds `valuePoints`,
`marketPrice`, or any engine stage — it is a second number shown next to the
board's own, for a human to weigh, exactly the same role `fp_tier` plays
next to the app's own computed tier.

FLUID BY DESIGN, NOT A DATA FILE: this module parses whatever copy of the
workbook the caller uploads, on demand, into THIS LEAGUE's
`settings.athleticProjections` — never a shared/global column, and nothing
here is ever written back into `fantasy_players`. The sheet is refreshed on
The Athletic's own schedule; the fix for a stale copy is re-uploading it,
the same discipline `AavPasteImport`/`YahooPasteImport` already follow for
their own "just copied this today" inputs. A raw uploaded workbook is never
persisted to disk or committed anywhere — it is parsed in memory and
discarded once matched.

COLUMNS ARE LOOKED UP BY HEADER TEXT, not fixed indices — same reasoning as
`projections.py`'s PROJ_SYN table and the offline
`data-pipeline/athletic_projections.py` twin this mirrors (kept independent
rather than imported: the backend has no dependency on data-pipeline, and
matching should not silently drift if that module's own norm() ever changes
for its own reasons).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import BinaryIO

from .base import NormPlayer

SUFFIX = re.compile(r"\b(jr|sr|ii|iii|iv|v)\b\.?", re.I)

# Header synonyms -> engine `proj` fields (points()/project_points() shape:
# passYd/passTD/int/rushYd/rushTD/rec/recYd/recTD). Matched case-insensitively
# against the workbook's own abbreviations, observed directly from real
# uploaded copies rather than guessed.
COL_SYN = {
    "passYd": ["payd"],
    "passTD": ["patd"],
    "int":    ["int"],
    "rushYd": ["ruyd"],
    "rushTD": ["rutd"],
    "rec":    ["rec"],
    "recYd":  ["rcyd"],
    "recTD":  ["rctd"],
}

# Sheets that carry per-player stat-line projections. DST/K carry no stat
# columns the app's points() formula scores — out of scope, same as the
# EXPERT_BLEND_W K/DST=1.0 (pure model, never touched) policy.
POS_SHEETS = ("QB", "RB", "WR", "TE")


def norm(n: str) -> str:
    n = (n or "").lower()
    n = re.sub(r"[.'`’]", "", n)
    n = SUFFIX.sub("", n)
    n = re.sub(r"[^a-z ]", " ", n)
    return re.sub(r"\s+", " ", n).strip()


@dataclass
class AthleticRow:
    name: str
    pos: str
    team: str
    proj: dict = field(default_factory=dict)


@dataclass
class AthleticUploadReport:
    rows: list[AthleticRow] = field(default_factory=list)
    sheets_found: list[str] = field(default_factory=list)


def _header_map(row) -> dict:
    out = {}
    for i, v in enumerate(row):
        if v is not None:
            out[str(v).strip().lower()] = i
    return out


def parse_workbook(fileobj: BinaryIO) -> AthleticUploadReport:
    """Parse an uploaded .xlsx into AthleticRow rows. Pure, no network/DB —
    `fileobj` is the upload's own in-memory/temp buffer, read once."""
    import openpyxl
    wb = openpyxl.load_workbook(fileobj, data_only=True, read_only=True)
    report = AthleticUploadReport()
    for sheet in POS_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        report.sheets_found.append(sheet)
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        try:
            header = _header_map(next(rows))
        except StopIteration:
            continue
        name_c = header.get("player")
        team_c = header.get("tm")
        if name_c is None:
            continue
        col_for = {}
        for field_name, syns in COL_SYN.items():
            for syn in syns:
                if syn in header:
                    col_for[field_name] = header[syn]
                    break
        for row in rows:
            if row is None or name_c >= len(row):
                continue
            name = row[name_c]
            if not name or not isinstance(name, str):
                continue
            proj = {}
            for field_name, idx in col_for.items():
                v = row[idx] if idx < len(row) else None
                if v is not None:
                    try:
                        proj[field_name] = float(v)
                    except (TypeError, ValueError):
                        pass
            team = row[team_c] if team_c is not None and team_c < len(row) else ""
            report.rows.append(AthleticRow(
                name=str(name).strip(), pos=sheet, team=str(team or "").strip().upper(),
                proj=proj,
            ))
    return report


def to_norm_players(report: AthleticUploadReport) -> list[NormPlayer]:
    """AthleticRow -> NormPlayer, so the existing `matching.py` index/matcher
    (the same one ESPN/Yahoo/AAV-paste import already go through) can be
    reused rather than writing a second name-matching path."""
    return [NormPlayer(name=r.name, pos=r.pos, team=r.team) for r in report.rows]
